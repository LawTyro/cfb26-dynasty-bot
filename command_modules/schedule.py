import os
import tempfile

import discord
from discord import app_commands
import openpyxl

import db
from utils import format_current_schedule_line, format_schedule_opponent, log_activity


def setup(tree, bot):
    schedule_group = app_commands.Group(name="schedule", description="Schedule commands")

    @schedule_group.command(name="import", description="Import schedule from Excel")
    @app_commands.checks.has_permissions(administrator=True)
    async def schedule_import(interaction: discord.Interaction, file: discord.Attachment):
        await interaction.response.defer(ephemeral=True)

        if not file.filename.lower().endswith(".xlsx"):
            return await interaction.followup.send(
                "Please upload an `.xlsx` Excel file.",
                ephemeral=True
            )

        temp_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        await file.save(temp_path)

        imported = 0
        skipped_aliases = set()

        try:
            workbook = openpyxl.load_workbook(temp_path, data_only=True)
            sheet = workbook.active

            db.clear_schedule()

            for row in range(2, sheet.max_row + 1):
                alias = sheet.cell(row=row, column=1).value

                if not alias:
                    continue

                alias = str(alias).strip()
                user_id = db.get_user_id_by_alias(alias)

                if not user_id:
                    skipped_aliases.add(alias)
                    continue

                for col in range(2, sheet.max_column + 1, 2):
                    week = sheet.cell(row=1, column=col).value
                    opponent = sheet.cell(row=row, column=col).value
                    user_marker = sheet.cell(row=row, column=col + 1).value

                    if not week or not opponent:
                        continue

                    is_user_game = (
                        str(user_marker).strip().upper() == "X"
                        if user_marker
                        else False
                    )

                    db.set_schedule_game(
                        user_id=user_id,
                        week=str(week).strip(),
                        opponent=str(opponent).strip(),
                        is_user_game=is_user_game
                    )

                    imported += 1

            message = f"✅ Imported **{imported}** scheduled game(s)."

            if skipped_aliases:
                message += "\n\n⚠️ These spreadsheet names do not have aliases:\n"
                message += ", ".join(sorted(skipped_aliases))

            await interaction.followup.send(message, ephemeral=True)

            log_description = f"Imported **{imported}** scheduled game(s) from `{file.filename}`."
            if skipped_aliases:
                log_description += "\nSkipped aliases: " + ", ".join(sorted(skipped_aliases))

            await log_activity(
                bot,
                interaction,
                "📅 Schedule Imported",
                log_description,
                discord.Color.green()
            )

        finally:
            try:
                os.remove(temp_path)
            except OSError:
                pass

    @schedule_group.command(name="clear", description="Clear imported schedule")
    @app_commands.checks.has_permissions(administrator=True)
    async def schedule_clear(interaction: discord.Interaction):
        db.clear_schedule()

        await interaction.response.send_message("🧹 Schedule cleared.", ephemeral=True)
        await log_activity(
            bot,
            interaction,
            "🧹 Schedule Cleared",
            "The imported schedule was cleared.",
            discord.Color.red()
        )

    @schedule_group.command(name="current", description="Show current week schedule")
    async def schedule_current(interaction: discord.Interaction):
        stage = db.get_setting("advance_stage", "")

        if not stage:
            return await interaction.response.send_message(
                "No current stage is set.",
                ephemeral=True
            )

        rows = db.get_schedule_for_week(stage)

        if not rows:
            return await interaction.response.send_message(
                f"No schedule found for **{stage}**.",
                ephemeral=True
            )

        lines = []

        for user_id, opponent, is_user_game in rows:
            member = interaction.guild.get_member(user_id)
            name = member.display_name if member else f"<@{user_id}>"
            lines.append(format_current_schedule_line(name, opponent, is_user_game))

        embed = discord.Embed(
            title=f"📅 {stage}",
            description="\n".join(lines),
            color=discord.Color.blue()
        )

        await interaction.response.send_message(embed=embed)

    @schedule_group.command(name="player", description="Show one player's schedule")
    async def schedule_player(interaction: discord.Interaction, member: discord.Member = None):
        target = member or interaction.user
        rows = db.get_schedule_for_player(target.id)

        if not rows:
            return await interaction.response.send_message(
                f"No schedule found for {target.display_name}.",
                ephemeral=True
            )

        current_stage = db.get_setting("advance_stage", "")
        lines = []

        for week, opponent, is_user_game in rows:
            display = format_schedule_opponent(opponent)

            short_week = week.replace("Week ", "W")
            line = f"{short_week:<3} {display}"

            if is_user_game:
                line += " *"

            if week == current_stage:
                line = f"> {line}"
            else:
                line = f"  {line}"

            lines.append(line)

        embed = discord.Embed(
            title=f"📅 {target.display_name}'s Schedule",
            description=(
                "```text\n"
                + "\n".join(lines)
                + "\n```"
                + "\n`* = user game`"
            ),
            color=discord.Color.blue()
        )

        await interaction.response.send_message(embed=embed, ephemeral=True)
        await log_activity(
            bot,
            interaction,
            "📅 Player Schedule Viewed",
            f"{interaction.user.display_name} viewed **{target.display_name}**'s schedule.",
            discord.Color.blue()
        )

    tree.add_command(schedule_group)
