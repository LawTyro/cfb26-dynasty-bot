import os
from dotenv import load_dotenv
load_dotenv()

import math
import asyncio
import tempfile
from datetime import datetime, timedelta, timezone

import discord
from discord.ext import commands
from discord import app_commands
import openpyxl

import db
import embeds

TOKEN = os.getenv("DISCORD_TOKEN")

intents = discord.Intents.default()
intents.members = True

bot = commands.Bot(command_prefix="!", intents=intents)
tree = bot.tree

ADVANCE_STAGES = [
    "Preseason",
    "Week 0",
    "Week 1",
    "Week 2",
    "Week 3",
    "Week 4",
    "Week 5",
    "Week 6",
    "Week 7",
    "Week 8",
    "Week 9",
    "Week 10",
    "Week 11",
    "Week 12",
    "Week 13",
    "Week 14",
    "Week 15",
    "Conference Championship",
    "Bowl Week 1",
    "Bowl Week 2",
    "Bowl Week 3",
    "Bowl Week 4",
    "Offseason Portal Week 1",
    "Offseason Portal Week 2",
    "Offseason Portal Week 3",
    "Offseason Portal Week 4",
]


async def stage_autocomplete(
    interaction: discord.Interaction,
    current: str
):
    current = current.lower()

    return [
        app_commands.Choice(name=stage, value=stage)
        for stage in ADVANCE_STAGES
        if current in stage.lower()
    ][:25]


def everyone_ready():
    players = db.get_players()

    if not players:
        return False

    ready_players = set(db.get_ready_players())

    return all(player in ready_players for player in players)


def get_unready_mentions():
    ready_players = set(db.get_ready_players())

    return [
        f"<@{uid}>"
        for uid in db.get_players()
        if uid not in ready_players
    ]


def get_output_channel(interaction=None):
    channel_id = db.get_setting("channel_id", "")

    if channel_id:
        channel = bot.get_channel(int(channel_id))

        if channel:
            return channel

    if interaction:
        return interaction.channel

    return None


def format_schedule_opponent(opponent):
    display = str(opponent).strip()

    if display.lower().startswith("at "):
        return "@ " + display[3:].strip()

    return display


def format_current_schedule_line(name, opponent, is_user_game):
    display = format_schedule_opponent(opponent)

    if display.startswith("@ "):
        line = f"🏈 {name} - {display}"
    else:
        line = f"🏈 {name} - vs {display}"

    if is_user_game:
        line += " *"

    return line


async def reminder_loop():
    await bot.wait_until_ready()

    while not bot.is_closed():
        try:
            remaining = embeds.get_remaining()

            if remaining:
                seconds = remaining.total_seconds()

                if seconds <= 0:
                    channel = get_output_channel()

                    if channel:
                        embed = discord.Embed(
                            title="🚨 Dynasty Is Advancing Now!",
                            description="The advance timer has expired.",
                            color=discord.Color.red(),
                            timestamp=datetime.now(timezone.utc)
                        )

                        await channel.send(
                            "@everyone",
                            embed=embed,
                            allowed_mentions=discord.AllowedMentions(everyone=True)
                        )

                    db.set_setting("advance_end", "")
                    db.clear_ready()
                    db.set_bool_setting("all_ready_sent", False)

                else:
                    days_left = math.ceil(seconds / 86400)
                    last_reminder_day = db.get_setting("last_reminder_day", "")

                    if str(days_left) != str(last_reminder_day):
                        db.set_setting("last_reminder_day", days_left)

                        channel = get_output_channel()

                        if channel:
                            unready = get_unready_mentions()

                            embed = discord.Embed(
                                title="🏈 Advance Reminder",
                                description=f"**{days_left} day(s) until advance**",
                                color=discord.Color.gold(),
                                timestamp=datetime.now(timezone.utc)
                            )

                            if unready:
                                embed.add_field(
                                    name="❌ Still Waiting On",
                                    value="\n".join(unready),
                                    inline=False
                                )

                                await channel.send(
                                    " ".join(unready),
                                    embed=embed,
                                    allowed_mentions=discord.AllowedMentions(users=True)
                                )
                            else:
                                embed.add_field(
                                    name="✅ Everyone Is Ready",
                                    value="No one is left to ready up.",
                                    inline=False
                                )

                                await channel.send(embed=embed)

        except Exception as e:
            print("Reminder error:", e)

        await asyncio.sleep(60)


player_group = app_commands.Group(name="player", description="Player management")


@player_group.command(name="add", description="Add player")
@app_commands.checks.has_permissions(administrator=True)
async def player_add(interaction: discord.Interaction, member: discord.Member):
    if db.is_player(member.id):
        return await interaction.response.send_message("Already added.", ephemeral=True)

    db.add_player(member.id)

    await interaction.response.send_message(f"✅ Added {member.display_name}")


@player_group.command(name="addall", description="Add all server members")
@app_commands.checks.has_permissions(administrator=True)
async def player_addall(interaction: discord.Interaction):
    added = 0
    skipped = 0

    for member in interaction.guild.members:
        if member.bot:
            skipped += 1
            continue

        if db.is_player(member.id):
            skipped += 1
            continue

        db.add_player(member.id)
        added += 1

    embed = discord.Embed(
        title="👥 Players Added",
        description=(
            f"✅ Added **{added}** player(s)\n"
            f"⏭️ Skipped **{skipped}** member(s)"
        ),
        color=discord.Color.green()
    )

    await interaction.response.send_message(embed=embed)


@player_group.command(name="remove", description="Remove player")
@app_commands.checks.has_permissions(administrator=True)
async def player_remove(interaction: discord.Interaction, member: discord.Member):
    if not db.is_player(member.id):
        return await interaction.response.send_message("Not found.", ephemeral=True)

    db.remove_player(member.id)

    await interaction.response.send_message(f"🛑 Removed {member.display_name}")


@player_group.command(name="alias", description="Set a player's schedule/import alias")
@app_commands.checks.has_permissions(administrator=True)
async def player_alias(
    interaction: discord.Interaction,
    member: discord.Member,
    alias: str
):
    if not db.is_player(member.id):
        return await interaction.response.send_message(
            "That user is not registered.",
            ephemeral=True
        )

    existing_user_id = db.get_user_id_by_alias(alias)

    if existing_user_id and existing_user_id != member.id:
        return await interaction.response.send_message(
            "That alias is already being used by another player.",
            ephemeral=True
        )

    db.set_player_alias(member.id, alias)

    await interaction.response.send_message(
        f"✅ Set alias for {member.display_name} to **{alias}**."
    )


@player_group.command(name="clearalias", description="Clear a player's alias")
@app_commands.checks.has_permissions(administrator=True)
async def player_clearalias(
    interaction: discord.Interaction,
    member: discord.Member
):
    if not db.is_player(member.id):
        return await interaction.response.send_message(
            "That user is not registered.",
            ephemeral=True
        )

    db.remove_player_alias(member.id)

    await interaction.response.send_message(
        f"🧹 Cleared alias for {member.display_name}."
    )


@player_group.command(name="list", description="List players")
async def player_list(interaction: discord.Interaction):
    ready_ids = set(db.get_ready_players())
    lines = []

    for uid in db.get_players():
        member = interaction.guild.get_member(uid)

        if not member:
            continue

        status = "✅" if uid in ready_ids else "❌"
        alias = db.get_player_alias(uid)

        if alias:
            lines.append(f"{status} {member.display_name} — alias: `{alias}`")
        else:
            lines.append(f"{status} {member.display_name}")

    embed = discord.Embed(
        title="👥 Dynasty Players",
        description="\n".join(lines) if lines else "No players.",
        color=discord.Color.blue()
    )

    await interaction.response.send_message(embed=embed)


tree.add_command(player_group)


history_group = app_commands.Group(name="history", description="Team history commands")


@history_group.command(name="view", description="View team history")
async def history_view(interaction: discord.Interaction, member: discord.Member = None):
    target = member or interaction.user
    teams = db.get_history(target.id)

    embed = discord.Embed(
        title=f"🏈 {target.display_name}'s Team History",
        color=discord.Color.blue()
    )

    embed.description = (
        "\n".join(f"- {team}" for team in teams)
        if teams
        else "No team history yet."
    )

    await interaction.response.send_message(embed=embed)


@history_group.command(name="all", description="Show all histories")
async def history_all(interaction: discord.Interaction):
    embed = discord.Embed(
        title="🏈 Team Histories",
        color=discord.Color.blue()
    )

    added = 0

    for uid in db.get_players():
        member = interaction.guild.get_member(uid)

        if not member:
            continue

        teams = db.get_history(uid)

        embed.add_field(
            name=member.display_name,
            value="\n".join(f"- {team}" for team in teams) if teams else "No team history",
            inline=False
        )

        added += 1

        if added >= 25:
            break

    if added == 0:
        embed.description = "No players found."

    await interaction.response.send_message(embed=embed)


@history_group.command(name="add", description="Add team history")
@app_commands.checks.has_permissions(administrator=True)
async def history_add(interaction: discord.Interaction, member: discord.Member, team: str):
    if not db.is_player(member.id):
        return await interaction.response.send_message("That player is not registered.", ephemeral=True)

    added = db.add_history(member.id, team)

    if not added:
        return await interaction.response.send_message(
            f"**{team}** already exists in {member.display_name}'s history.",
            ephemeral=True
        )

    await interaction.response.send_message(
        f"🏈 Added **{team}** to {member.display_name}'s history."
    )


@history_group.command(name="remove", description="Remove team history")
@app_commands.checks.has_permissions(administrator=True)
async def history_remove(interaction: discord.Interaction, member: discord.Member, team: str):
    removed = db.remove_history(member.id, team)

    if not removed:
        return await interaction.response.send_message(f"**{team}** not found.", ephemeral=True)

    await interaction.response.send_message(
        f"🧹 Removed **{removed}** from {member.display_name}'s history."
    )


@history_group.command(name="reset", description="Reset history")
@app_commands.checks.has_permissions(administrator=True)
async def history_reset(
    interaction: discord.Interaction,
    member: discord.Member = None,
    all_players: bool = False
):
    if all_players:
        db.reset_history()
        return await interaction.response.send_message("🧹 Reset history for all players.")

    target = member or interaction.user
    db.reset_history(target.id)

    await interaction.response.send_message(f"🧹 Reset history for {target.display_name}.")


tree.add_command(history_group)


h2h_group = app_commands.Group(name="head2head", description="Head-to-head records")


@h2h_group.command(name="add", description="Add a head-to-head result")
@app_commands.checks.has_permissions(administrator=True)
async def h2h_add(interaction: discord.Interaction, winner: discord.Member, loser: discord.Member):
    if winner.id == loser.id:
        return await interaction.response.send_message("Winner and loser cannot be the same.", ephemeral=True)

    if not db.is_player(winner.id) or not db.is_player(loser.id):
        return await interaction.response.send_message("Both users must be registered players.", ephemeral=True)

    db.add_h2h_game(winner.id, loser.id)

    embed = discord.Embed(
        title="🏈 Head-to-Head Result Added",
        description=f"**{winner.display_name}** defeated **{loser.display_name}**.",
        color=discord.Color.green()
    )

    await interaction.response.send_message(embed=embed)


@h2h_group.command(name="remove", description="Remove latest matching head-to-head result")
@app_commands.checks.has_permissions(administrator=True)
async def h2h_remove(interaction: discord.Interaction, winner: discord.Member, loser: discord.Member):
    removed = db.remove_latest_h2h_game(winner.id, loser.id)

    if not removed:
        return await interaction.response.send_message(
            f"No matching result found for {winner.display_name} over {loser.display_name}.",
            ephemeral=True
        )

    embed = discord.Embed(
        title="🧹 Head-to-Head Result Removed",
        description=f"Removed latest result: **{winner.display_name}** defeated **{loser.display_name}**.",
        color=discord.Color.red()
    )

    await interaction.response.send_message(embed=embed)


@h2h_group.command(name="reset", description="Reset head-to-head records")
@app_commands.checks.has_permissions(administrator=True)
async def h2h_reset(
    interaction: discord.Interaction,
    player1: discord.Member = None,
    player2: discord.Member = None,
    all_players: bool = False
):
    if all_players:
        db.reset_h2h()

        embed = discord.Embed(
            title="🧹 Head-to-Head Records Reset",
            description="All head-to-head records have been deleted.",
            color=discord.Color.red()
        )

        return await interaction.response.send_message(embed=embed)

    if not player1 or not player2:
        return await interaction.response.send_message(
            "Provide both players or set all_players to True.",
            ephemeral=True
        )

    db.reset_h2h(player1.id, player2.id)

    embed = discord.Embed(
        title="🧹 Matchup Reset",
        description=f"Removed all games between **{player1.display_name}** and **{player2.display_name}**.",
        color=discord.Color.red()
    )

    await interaction.response.send_message(embed=embed)


@h2h_group.command(name="view", description="View record between two players")
async def h2h_view(interaction: discord.Interaction, player1: discord.Member, player2: discord.Member):
    if player1.id == player2.id:
        return await interaction.response.send_message("Choose two different players.", ephemeral=True)

    await interaction.response.send_message(
        embed=embeds.make_h2h_view_embed(player1, player2)
    )


@h2h_group.command(name="player", description="View one player's head-to-head records")
async def h2h_player(interaction: discord.Interaction, player: discord.Member = None):
    target = player or interaction.user

    await interaction.response.send_message(
        embed=embeds.make_h2h_player_embed(interaction.guild, target)
    )


@h2h_group.command(name="standings", description="View overall head-to-head standings")
async def h2h_standings(interaction: discord.Interaction):
    await interaction.response.send_message(
        embed=embeds.make_h2h_standings_embed(interaction.guild)
    )


tree.add_command(h2h_group)


schedule_group = app_commands.Group(name="schedule", description="Schedule commands")


@schedule_group.command(name="import", description="Import schedule from Excel")
@app_commands.checks.has_permissions(administrator=True)
async def schedule_import(
    interaction: discord.Interaction,
    file: discord.Attachment
):
    await interaction.response.defer(ephemeral=True)

    if not file.filename.lower().endswith(".xlsx"):
        return await interaction.followup.send(
            "Please upload an `.xlsx` Excel file.",
            ephemeral=True
        )

    temp_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    await file.save(temp_path)

    imported = 0
    skipped_aliases = set()

    try:
        workbook = openpyxl.load_workbook(temp_path, data_only=True)
        sheet = workbook.active

        db.clear_schedule()

        for row in range(2, sheet.max_row + 1):
            alias = sheet.cell(row=row, column=1).value

            if not alias:
                continue

            alias = str(alias).strip()
            user_id = db.get_user_id_by_alias(alias)

            if not user_id:
                skipped_aliases.add(alias)
                continue

            for col in range(2, sheet.max_column + 1, 2):
                week = sheet.cell(row=1, column=col).value
                opponent = sheet.cell(row=row, column=col).value
                user_marker = sheet.cell(row=row, column=col + 1).value

                if not week or not opponent:
                    continue

                is_user_game = (
                    str(user_marker).strip().upper() == "X"
                    if user_marker
                    else False
                )

                db.set_schedule_game(
                    user_id=user_id,
                    week=str(week).strip(),
                    opponent=str(opponent).strip(),
                    is_user_game=is_user_game
                )

                imported += 1

        message = f"✅ Imported **{imported}** scheduled game(s)."

        if skipped_aliases:
            message += "\n\n⚠️ These spreadsheet names do not have aliases:\n"
            message += ", ".join(sorted(skipped_aliases))

        await interaction.followup.send(message, ephemeral=True)

    finally:
        try:
            os.remove(temp_path)
        except OSError:
            pass


@schedule_group.command(name="clear", description="Clear imported schedule")
@app_commands.checks.has_permissions(administrator=True)
async def schedule_clear(interaction: discord.Interaction):
    db.clear_schedule()
    await interaction.response.send_message("🧹 Schedule cleared.", ephemeral=True)


@schedule_group.command(name="current", description="Show current week schedule")
async def schedule_current(interaction: discord.Interaction):
    stage = db.get_setting("advance_stage", "")

    if not stage:
        return await interaction.response.send_message(
            "No current stage is set.",
            ephemeral=True
        )

    rows = db.get_schedule_for_week(stage)

    if not rows:
        return await interaction.response.send_message(
            f"No schedule found for **{stage}**.",
            ephemeral=True
        )

    lines = []

    for user_id, opponent, is_user_game in rows:
        member = interaction.guild.get_member(user_id)
        name = member.display_name if member else f"<@{user_id}>"
        lines.append(format_current_schedule_line(name, opponent, is_user_game))

    embed = discord.Embed(
        title=f"📅 {stage}",
        description="\n".join(lines),
        color=discord.Color.blue()
    )

    await interaction.response.send_message(embed=embed)


@schedule_group.command(name="player", description="Show one player's schedule")
async def schedule_player(
    interaction: discord.Interaction,
    member: discord.Member = None
):
    target = member or interaction.user
    rows = db.get_schedule_for_player(target.id)

    if not rows:
        return await interaction.response.send_message(
            f"No schedule found for {target.display_name}.",
            ephemeral=True
        )

    current_stage = db.get_setting("advance_stage", "")
    lines = []

for week, opponent, is_user_game in rows:
    display = format_schedule_opponent(opponent)

    line = f"{week:<8} | {display}"

    if is_user_game:
        line += " *"

    if week == current_stage:
        line = f"> {line}"
    else:
        line = f"  {line}"

    lines.append(line)

    embed = discord.Embed(
    title=f"📅 {target.display_name}'s Schedule",
    description=(
        "```text\n"
        + "\n".join(lines)
        + "\n```"
        + "\n`* = user game`"
    ),
    color=discord.Color.blue()
)

    await interaction.response.send_message(embed=embed)


tree.add_command(schedule_group)


@tree.command(name="setchannel", description="Set bot announcement channel")
@app_commands.checks.has_permissions(administrator=True)
async def setchannel(
    interaction: discord.Interaction,
    channel: discord.TextChannel
):
    db.set_setting("channel_id", channel.id)

    await interaction.response.send_message(
        f"✅ Bot announcements will go to {channel.mention}.",
        ephemeral=True
    )


@tree.command(name="advance", description="Start/reset advance timer")
@app_commands.autocomplete(stage=stage_autocomplete)
async def advance(
    interaction: discord.Interaction,
    stage: str = None
):
    days = int(db.get_setting("advance_days", "4"))
    new_end = datetime.now(timezone.utc) + timedelta(days=days)
    selected_stage = stage or ""

    db.set_setting("advance_end", new_end.isoformat())
    db.set_setting("last_reminder_day", "")
    db.set_bool_setting("all_ready_sent", False)
    db.set_setting("advance_stage", selected_stage)
    db.clear_ready()

    target_channel = get_output_channel(interaction)

    description = f"Advance is in **{days} day(s)**."

    if selected_stage:
        description += f"\nAdvanced to: **{selected_stage}**"

    embed = discord.Embed(
        title="🏈 Advance Timer Started",
        description=description,
        color=discord.Color.gold(),
        timestamp=datetime.now(timezone.utc)
    )

    await target_channel.send(
        "@everyone",
        embed=embed,
        allowed_mentions=discord.AllowedMentions(everyone=True)
    )

    await interaction.response.send_message(
        f"✅ Advance started in {target_channel.mention}.",
        ephemeral=True
    )


@tree.command(name="cancel", description="Cancel advance")
@app_commands.checks.has_permissions(administrator=True)
async def cancel(interaction: discord.Interaction):
    db.set_setting("advance_end", "")
    db.clear_ready()
    db.set_bool_setting("all_ready_sent", False)

    target_channel = get_output_channel(interaction)

    embed = discord.Embed(
        title="🛑 Advance Cancelled",
        description="The current advance timer has been cancelled.",
        color=discord.Color.red()
    )

    await target_channel.send(embed=embed)

    await interaction.response.send_message(
        f"✅ Advance cancellation posted in {target_channel.mention}.",
        ephemeral=True
    )


@tree.command(name="extend", description="Extend timer")
@app_commands.checks.has_permissions(administrator=True)
async def extend(interaction: discord.Interaction, days: int):
    remaining = embeds.get_remaining()

    if not remaining:
        return await interaction.response.send_message("No active advance.", ephemeral=True)

    new_end = datetime.now(timezone.utc) + remaining + timedelta(days=days)

    db.set_setting("advance_end", new_end.isoformat())
    db.set_setting("last_reminder_day", "")

    target_channel = get_output_channel(interaction)

    embed = discord.Embed(
        title="⏳ Advance Timer Extended",
        description=f"Extended by **{days} day(s)**.",
        color=discord.Color.gold()
    )

    await target_channel.send(embed=embed)

    await interaction.response.send_message(
        f"✅ Extension posted in {target_channel.mention}.",
        ephemeral=True
    )


@tree.command(name="setdays", description="Set default advance days")
@app_commands.checks.has_permissions(administrator=True)
async def setdays(interaction: discord.Interaction, days: int):
    if days <= 0:
        return await interaction.response.send_message("Days must be greater than 0.", ephemeral=True)

    db.set_setting("advance_days", days)

    embed = discord.Embed(
        title="✅ Default Advance Length Updated",
        description=f"Default advance length set to **{days} day(s)**.",
        color=discord.Color.green()
    )

    await interaction.response.send_message(embed=embed, ephemeral=True)


@tree.command(name="ready", description="Mark ready")
async def ready(interaction: discord.Interaction):
    uid = interaction.user.id

    if not db.is_player(uid):
        return await interaction.response.send_message("Not registered.", ephemeral=True)

    if uid in db.get_ready_players():
        return await interaction.response.send_message("Already ready.", ephemeral=True)

    db.mark_ready(uid)

    await interaction.response.send_message("✅ Ready!")

    if db.get_setting("advance_end", "") and everyone_ready() and not db.get_bool_setting("all_ready_sent"):
        db.set_bool_setting("all_ready_sent", True)

        channel = get_output_channel(interaction)

        if channel:
            commissioner_role = discord.utils.get(channel.guild.roles, name="Commissioner")
            commissioner_ping = commissioner_role.mention if commissioner_role else "@Commissioner"

            embed = discord.Embed(
                title="🏈 Everyone Is Ready!",
                description="Commissioners can advance the league.",
                color=discord.Color.green(),
                timestamp=datetime.now(timezone.utc)
            )

            await channel.send(
                commissioner_ping,
                embed=embed,
                allowed_mentions=discord.AllowedMentions(roles=True)
            )
            
@tree.command(name="setready", description="Mark another player as ready")
@app_commands.checks.has_permissions(administrator=True)
async def setready(interaction: discord.Interaction, member: discord.Member):
    if not db.is_player(member.id):
        return await interaction.response.send_message(
            "That user is not registered.",
            ephemeral=True
        )

    if member.id in db.get_ready_players():
        return await interaction.response.send_message(
            f"{member.display_name} is already ready.",
            ephemeral=True
        )

    db.mark_ready(member.id)

    await interaction.response.send_message(
        f"✅ Marked {member.display_name} as ready.",
        ephemeral=True
    )

    if db.get_setting("advance_end", "") and everyone_ready() and not db.get_bool_setting("all_ready_sent"):
        db.set_bool_setting("all_ready_sent", True)

        channel = get_output_channel(interaction)

        if channel:
            commissioner_role = discord.utils.get(channel.guild.roles, name="Commissioner")
            commissioner_ping = commissioner_role.mention if commissioner_role else "@Commissioner"

            embed = discord.Embed(
                title="🏈 Everyone Is Ready!",
                description="Commissioners can advance the league.",
                color=discord.Color.green(),
                timestamp=datetime.now(timezone.utc)
            )

            await channel.send(
                commissioner_ping,
                embed=embed,
                allowed_mentions=discord.AllowedMentions(roles=True)
            )

@tree.command(name="unready", description="Mark unready")
async def unready(interaction: discord.Interaction):
    db.mark_unready(interaction.user.id)
    db.set_bool_setting("all_ready_sent", False)

    await interaction.response.send_message("↩️ Unready")


@tree.command(name="clearready", description="Clear all ready statuses")
@app_commands.checks.has_permissions(administrator=True)
async def clearready(interaction: discord.Interaction):
    db.clear_ready()
    db.set_bool_setting("all_ready_sent", False)

    target_channel = get_output_channel(interaction)

    embed = discord.Embed(
        title="🧹 Ready Status Cleared",
        description="All players have been marked not ready.",
        color=discord.Color.red()
    )

    await target_channel.send(embed=embed)

    await interaction.response.send_message(
        f"✅ Ready statuses cleared and posted in {target_channel.mention}.",
        ephemeral=True
    )


@tree.command(name="status", description="Show dynasty status")
async def status(interaction: discord.Interaction):
    await interaction.response.send_message(
        embed=embeds.make_status_embed(interaction.guild)
    )


@tree.command(name="help", description="Show commands")
async def help_command(interaction: discord.Interaction):
    await interaction.response.send_message(
        embed=embeds.make_help_embed(),
        ephemeral=True
    )


@bot.event
async def on_ready():
    db.init_db()
    db.migrate_json_if_needed()

    synced = await tree.sync()
    print(f"Synced {len(synced)} command(s)")
    print(f"Commands: {[cmd.name for cmd in synced]}")
    print(f"Logged in as {bot.user}")

    if not hasattr(bot, "reminder_task"):
        bot.reminder_task = asyncio.create_task(reminder_loop())


bot.run(TOKEN)
